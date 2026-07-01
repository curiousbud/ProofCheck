"""Human-readable xlsx report generated from a :class:`RunResult` (openpyxl).

A Summary sheet with a plain-English overview and friendly counts, plus one sheet per
checked column whose rows read in everyday language ("Found", "Found with differences",
"Not found", "Blank") with the exact text we found. Status cells are color-coded to match
the HTML report and web UI.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import humanize
from .models import RunResult, Status


def _clean(value):
    """Strip control chars openpyxl rejects (OCR text can contain them); pass-through others.

    Excel/openpyxl forbid most C0 control characters in cell text (``\\x00-\\x08``,
    ``\\x0b``, ``\\x0c``, ``\\x0e-\\x1f``); tab/newline/carriage-return are fine. OCR output
    occasionally emits one of the forbidden ones, which otherwise crashes the whole report
    with ``IllegalCharacterError``. Only strings are touched.
    """
    return ILLEGAL_CHARACTERS_RE.sub("", value) if isinstance(value, str) else value


def _append(ws, values) -> None:
    """``ws.append`` with every string cell sanitized against illegal control chars."""
    ws.append([_clean(v) for v in values])

# Solid fills matching the shared status palette.
_FILLS = {
    Status.EXACT: PatternFill("solid", fgColor="2E7D32"),
    Status.FUZZY: PatternFill("solid", fgColor="F59E0B"),
    Status.MISSING: PatternFill("solid", fgColor="C62828"),
    Status.SKIPPED: PatternFill("solid", fgColor="9E9E9E"),
}
_WHITE = Font(color="FFFFFF", bold=True)
_HEADER = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _autosize(ws, max_width: int = 70) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                # Use the longest line so wrapped detail cells don't force huge columns.
                longest = max((len(line) for line in str(cell.value).splitlines()), default=0)
                widths[cell.column] = min(max(widths.get(cell.column, 0), longest), max_width)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width + 2


def build(result: RunResult) -> Workbook:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    s = result.summary
    summary_ws.append(["ProofCheck results"])
    summary_ws["A1"].font = Font(bold=True, size=14)
    _append(summary_ws, [humanize.summary_sentence(result)])
    summary_ws.append([])
    rows = [
        ("Spreadsheet", result.meta.excel),
        ("PDF", result.meta.pdf),
        ("Checked on", result.meta.timestamp),
        ("", ""),
        ("Values checked", s.total - s.skipped),
        ("Found", s.exact),
        ("Found with differences", s.fuzzy),
        ("Not found", s.missing),
        ("Blank (skipped)", s.skipped),
        ("Match rate", f"{s.pass_rate * 100:.0f}%"),
    ]
    for label_text, value in rows:
        _append(summary_ws, [label_text, value])
        if label_text:
            summary_ws.cell(row=summary_ws.max_row, column=1).font = _HEADER

    if result.warnings:
        summary_ws.append([])
        summary_ws.append(["Notes"])
        summary_ws.cell(row=summary_ws.max_row, column=1).font = _HEADER
        for w in result.warnings:
            _append(summary_ws, ["", w])
    _autosize(summary_ws)

    for col in result.columns:
        # Sheet titles are capped at 31 chars and can't contain certain characters.
        title = col.name[:31] or "Column"
        ws = wb.create_sheet(title=title)
        ws.append(["Row", "Value in your spreadsheet", "Result", "Matched via", "Details"])
        for cell in ws[1]:
            cell.font = _HEADER
        for r in col.results:
            _append(ws, [
                r.row,
                r.expected,
                f"{humanize.icon(r.status)} {humanize.label(r.status)}",
                humanize.source_label(r.source),
                humanize.detail(r),
            ])
            row_idx = ws.max_row
            result_cell = ws.cell(row=row_idx, column=3)
            result_cell.fill = _FILLS[r.status]
            result_cell.font = _WHITE
            ws.cell(row=row_idx, column=5).alignment = _WRAP
        _autosize(ws)

    return wb


def write(result: RunResult, path: str) -> None:
    build(result).save(path)

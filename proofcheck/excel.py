"""Excel loading and inspection (openpyxl).

Read-only access to .xlsx/.xlsm workbooks: list sheets, read header rows, and pull
the values of selected columns. No matching logic here.
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl import load_workbook


class ExcelError(Exception):
    """Raised for user-facing spreadsheet problems (missing sheet/column, bad file)."""


@dataclass
class ColumnData:
    """Values of one column, paired with their 1-based spreadsheet row numbers."""

    name: str
    # (row_number, raw_value) for every data row below the header.
    cells: list[tuple[int, object]]


def _open(path: str):
    try:
        # read_only is fast and memory-light; data_only returns computed values, not formulae.
        return load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a grab-bag of exceptions on bad files
        raise ExcelError(f"Could not open Excel file: {exc}") from exc


def inspect(path: str) -> dict[str, list[str]]:
    """Return {sheet_name: [header, ...]} using row 1 as headers for every sheet.

    Powers the web column picker so users don't have to type exact header names.
    """
    wb = _open(path)
    try:
        result: dict[str, list[str]] = {}
        for ws in wb.worksheets:
            headers = _read_header(ws, header_row=1)
            result[ws.title] = headers
        return result
    finally:
        wb.close()


def sheet_names(path: str) -> list[str]:
    wb = _open(path)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _read_header(ws, header_row: int) -> list[str]:
    """Read a header row into a list of stripped strings (blank cells -> '')."""
    rows = ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    try:
        raw = next(rows)
    except StopIteration:
        return []
    return ["" if v is None else str(v).strip() for v in raw]


def load_columns(
    path: str,
    *,
    sheet: str | None = None,
    header_row: int = 1,
    columns: list[str] | None = None,
    all_columns: bool = False,
) -> list[ColumnData]:
    """Load the requested columns' values keyed by spreadsheet row number.

    Selects ``sheet`` (or the active sheet when None), reads headers from
    ``header_row``, then returns one ``ColumnData`` per requested column.
    """
    wb = _open(path)
    try:
        if sheet is None:
            ws = wb.active
        elif sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            raise ExcelError(
                f"Sheet {sheet!r} not found. Available: {', '.join(wb.sheetnames)}"
            )

        headers = _read_header(ws, header_row)
        if not headers:
            raise ExcelError(f"No header row found at row {header_row}.")

        # Map header name -> column index. First occurrence wins on duplicates.
        header_index: dict[str, int] = {}
        for idx, name in enumerate(headers):
            if name and name not in header_index:
                header_index[name] = idx

        if all_columns:
            selected = [h for h in headers if h]
        else:
            selected = columns or []
            missing = [c for c in selected if c not in header_index]
            if missing:
                raise ExcelError(
                    f"Column(s) not found: {', '.join(missing)}. "
                    f"Available: {', '.join(header_index)}"
                )

        result = [ColumnData(name=name, cells=[]) for name in selected]
        by_name = {cd.name: cd for cd in result}

        for row_num, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            for name in selected:
                col_idx = header_index[name]
                value = row[col_idx] if col_idx < len(row) else None
                by_name[name].cells.append((row_num, value))

        return result
    finally:
        wb.close()
